import sys
import os
import json
import math
import re
import ezdxf
import openpyxl
from openpyxl.cell.cell import MergedCell

from ezdxf.recover import readfile as recover_readfile
from ezdxf.addons.drawing.pyqt import PyQtBackend
from ezdxf.math import Vec3
from ezdxf.layouts import Layout
from ezdxf.addons.drawing.frontend import Frontend
from ezdxf.addons.drawing.properties import Properties, RenderContext
from ezdxf.entities import DXFGraphic
import ezdxf.bbox

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QGraphicsScene, QGraphicsView, QMessageBox,
    QFileDialog, QWidget, QSplitter, QVBoxLayout, QGridLayout, QLabel,
    QLineEdit, QComboBox, QFrame, QHBoxLayout, QDateEdit, QStyleFactory,
    QPushButton, QAction, QToolBar
)
from PyQt5.QtGui import (
    QColor, QWheelEvent, QFont, QDragEnterEvent, QDropEvent, QPixmap, QScreen,
    QIcon, QPainter, QPen, QBrush
)
from PyQt5.QtCore import Qt, QPoint, QPointF, pyqtSignal as Signal, QDate

from qt_material import apply_stylesheet

# ==============================================================================
#      2. KONSTANTEN UND HILFSKLASSEN
# ==============================================================================

ZOOM_FACTOR = 1.15


def resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class ClickableLineEdit(QLineEdit):
    clicked = Signal()

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)


class IsoFitsCalculator:
    # ... (unverändert) ...
    def __init__(self, data_folder_path):
        self.tolerances_data = []
        self.available_fits = [""]
        self._load_data(data_folder_path)

    def _load_data(self, data_folder_path):
        tolerances_path = os.path.join(data_folder_path, "tolerances.json")
        try:
            with open(tolerances_path, 'r', encoding='utf-8') as f:
                self.tolerances_data = json.load(f)
            all_fits = set(entry["toleranzklasse"] for entry in self.tolerances_data)
            self.available_fits.extend(sorted(list(all_fits)))
        except (FileNotFoundError, json.JSONDecodeError) as e:
            QMessageBox.critical(None, "Fataler Fehler", f"Datei 'tolerances.json' nicht gefunden/lesbar: {e}")
            sys.exit(1)

    def calculate(self, nominal_size, fit_string):
        if not self.tolerances_data:
            return None
        try:
            for entry in self.tolerances_data:
                if (entry["toleranzklasse"].lower() == fit_string.lower() and
                        entry["lowerlimit"] < nominal_size <= entry["upperlimit"]):
                    return entry["es"] / 1000.0, entry["ei"] / 1000.0
            return None
        except Exception:
            return None


class PatchedPyQtBackend(PyQtBackend):
    def _get_pen(self, properties: Properties) -> QPen:
        pen = super()._get_pen(properties)
        pen.setColor(QColor(0, 0, 0))
        return pen

    def _get_brush(self, properties: Properties) -> QBrush:
        brush = super()._get_brush(properties)
        brush.setColor(QColor(0, 0, 0))
        return brush


# ==============================================================================
#      3. HAUPT-WIDGETS
# ==============================================================================

class DXFWidget(QWidget):
    dimension_clicked = Signal(str)
    text_clicked = Signal(str)
    CLICK_RADIUS_PIXELS = 50

    def __init__(self, parent=None):
        super().__init__(parent)
        self.doc, self.active_layout = None, None
        self.selection_mode = 'dimension'
        self.scene = QGraphicsScene()
        self.view = QGraphicsView(self.scene)
        self.view.setRenderHint(QPainter.Antialiasing)
        self.view.setAcceptDrops(False)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.view)
        self.setLayout(layout)
        self.view.scale(1, -1)
        self.view.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorViewCenter)
        self.view.setDragMode(QGraphicsView.DragMode.NoDrag)
        self._is_panning = False
        self._pan_start_pos = QPointF()
        self.view.mousePressEvent = self.handle_mouse_press
        self.view.mouseMoveEvent = self.handle_mouse_move
        self.view.mouseReleaseEvent = self.handle_mouse_release
        self.view.wheelEvent = self.handle_wheel_event

    def set_selection_mode(self, mode: str):
        if mode in ['dimension', 'text']:
            self.selection_mode = mode
        else:
            print(f"WARNUNG: Unbekannter Auswahlmodus '{mode}'. Ignoriert.")

    def handle_wheel_event(self, event: QWheelEvent):
        anchor_point = self.view.mapToScene(event.pos())
        factor = ZOOM_FACTOR if event.angleDelta().y() > 0 else 1 / ZOOM_FACTOR
        self.view.translate(anchor_point.x(), anchor_point.y())
        self.view.scale(factor, factor)
        self.view.translate(-anchor_point.x(), -anchor_point.y())

    def load_dxf(self, filepath):
        try:
            self.doc, auditor = recover_readfile(filepath)
            self.draw_dxf()
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"DXF-Ladefehler: {e}")

    def _get_layout_to_display(self) -> Layout | None:
        if not self.doc:
            return None
        msp = self.doc.modelspace()
        if len(msp) > 0:
            return msp
        for layout in self.doc.layouts:
            if not layout.is_modelspace and len(layout) > 0:
                return layout
        return msp

    def draw_dxf(self):
        self.scene.clear()
        self.view.setBackgroundBrush(QColor(240, 240, 240))
        if not self.doc:
            return
        self.active_layout = self._get_layout_to_display()
        if not self.active_layout:
            print("Kein darstellbares Layout gefunden.")
            return
        try:
            backend = PatchedPyQtBackend(self.scene)
            backend.lineweight_scaling = 5.0
            ctx = RenderContext(self.doc)
            frontend = Frontend(ctx, backend)
            frontend.draw_layout(self.active_layout, finalize=True)
            self.view.fitInView(self.scene.itemsBoundingRect(), Qt.KeepAspectRatio)
        except Exception as e:
            print(f"Zeichnen fehlgeschlagen: {e}")

    def handle_mouse_press(self, event):
        if event.button() == Qt.MouseButton.RightButton:
            self._is_panning = True
            self._pan_start_pos = event.pos()
            self.view.setCursor(Qt.CursorShape.ClosedHandCursor)
            event.accept()
        elif event.button() == Qt.MouseButton.LeftButton:
            if not self.active_layout:
                super(QGraphicsView, self.view).mousePressEvent(event)
                return
            scene_pos = self.view.mapToScene(event.pos())
            world_pos = Vec3(scene_pos.x(), scene_pos.y())
            p2_pos = event.pos() + QPointF(self.CLICK_RADIUS_PIXELS, 0)
            p2 = self.view.mapToScene(p2_pos.toPoint())
            search_radius = abs(p2.x() - scene_pos.x())
            self.inspect_entity_at(world_pos, search_radius)
            if self.selection_mode == 'dimension':
                self.find_closest_dimension(world_pos, search_radius)
            elif self.selection_mode == 'text':
                self.find_closest_text(world_pos, search_radius)
            event.accept()
        else:
            super(QGraphicsView, self.view).mousePressEvent(event)

    def handle_mouse_move(self, event):
        if self._is_panning:
            delta = event.pos() - self._pan_start_pos
            self._pan_start_pos = event.pos()
            h_bar = self.view.horizontalScrollBar()
            v_bar = self.view.verticalScrollBar()
            h_bar.setValue(h_bar.value() - int(delta.x()))
            v_bar.setValue(v_bar.value() - int(delta.y()))
            event.accept()
        else:
            super(QGraphicsView, self.view).mouseMoveEvent(event)

    def handle_mouse_release(self, event):
        if event.button() == Qt.MouseButton.RightButton and self._is_panning:
            self._is_panning = False
            self.view.setCursor(Qt.CursorShape.ArrowCursor)
            event.accept()
        else:
            super(QGraphicsView, self.view).mouseReleaseEvent(event)

    def inspect_entity_at(self, world_pos: Vec3, radius: float):
        if not self.active_layout: return

        def find_closest_recursive(entities):
            closest = None
            min_dist = float('inf')
            for entity in entities:
                try:
                    if entity.dxftype() == 'INSERT':
                        sub_entity, dist = find_closest_recursive(entity.virtual_entities())
                        if sub_entity and dist < min_dist:
                            min_dist = dist
                            closest = sub_entity
                    else:
                        bbox = ezdxf.bbox.extents([entity], fast=True)
                        if bbox.has_data:
                            dist = world_pos.distance(bbox.center)
                            if dist < min_dist:
                                min_dist = dist
                                closest = entity
                except (RuntimeError, TypeError):
                    continue
            return closest, min_dist

        closest_entity, dist = find_closest_recursive(self.active_layout)
        if closest_entity and dist < radius:
            print("\n" + "=" * 20 + " Entity Inspector " + "=" * 20)
            print(f"DXF Type: {closest_entity.dxftype()}")
            print(f"Handle:   {closest_entity.dxf.handle}")
            print(f"Layer:    {closest_entity.dxf.layer}")
            print(f"Color:    {closest_entity.dxf.color} (256=ByLayer, 0=ByBlock)")
            print("-" * 58)
            print("Alle DXF-Attribute:")
            for key, value in closest_entity.dxf.all_existing_dxf_attribs().items():
                print(f"  - {key}: {value}")
            print("=" * 58 + "\n")

    def find_closest_dimension(self, world_pos, radius):
        if not self.active_layout: return

        def get_dimension_value(dim) -> float | None:
            try:
                measurement = dim.get_measurement()
                if isinstance(measurement, (int, float)): return float(measurement)
            except (TypeError, ValueError):
                pass
            user_text = dim.dxf.text
            if user_text and user_text != "<>":
                cleaned_text = user_text.replace(',', '.')
                match = re.search(r'[-+]?\d*\.?\d+', cleaned_text)
                if match:
                    try:
                        return float(match.group(0))
                    except (ValueError, TypeError):
                        pass
            try:
                for primitive in dim.virtual_entities():
                    if primitive.dxftype() in {'TEXT', 'MTEXT'}:
                        text_content = primitive.dxf.text if primitive.dxftype() == 'TEXT' else primitive.plain_text()
                        cleaned_text = text_content.replace(',', '.')
                        match = re.search(r'[-+]?\d*\.?\d+', cleaned_text)
                        if match: return float(match.group(0))
            except Exception:
                pass
            return None

        def find_dims_recursive(entities):
            found = []
            for entity in entities:
                try:
                    if entity.dxftype() == 'DIMENSION':
                        dist = float('inf')
                        text_entity = None
                        for sub in entity.virtual_entities():
                            if sub.dxftype() in {'TEXT', 'MTEXT'}:
                                text_entity = sub
                                break
                        if text_entity and text_entity.dxf.hasattr('insert'):
                            dist = world_pos.distance(text_entity.dxf.insert)
                        else:
                            bbox = ezdxf.bbox.extents([entity], fast=True)
                            if bbox.has_data:
                                dist = world_pos.distance(bbox.center)
                        if dist < radius:
                            value = get_dimension_value(entity)
                            if value is not None:
                                found.append((dist, value))
                    elif entity.dxftype() == 'INSERT':
                        found.extend(find_dims_recursive(entity.virtual_entities()))
                except (RuntimeError, TypeError):
                    continue
            return found

        found_dimensions = find_dims_recursive(self.active_layout)
        if found_dimensions:
            found_dimensions.sort(key=lambda x: x[0])
            self.dimension_clicked.emit(f"{found_dimensions[0][1]:.4f}")

    def find_closest_text(self, world_pos, radius):
        if not self.active_layout: return

        def find_texts_recursive(entities):
            found = []
            for entity in entities:
                try:
                    if entity.dxftype() in {'TEXT', 'MTEXT'}:
                        if entity.dxf.hasattr('insert'):
                            dist = world_pos.distance(entity.dxf.insert)
                            if dist < radius:
                                found.append((dist, entity))
                    elif entity.dxftype() == 'INSERT':
                        found.extend(find_texts_recursive(entity.virtual_entities()))
                except (RuntimeError, TypeError):
                    continue
            return found

        found_texts = find_texts_recursive(self.active_layout)
        if found_texts:
            found_texts.sort(key=lambda x: x[0])
            closest_entity = found_texts[0][1]
            text = closest_entity.dxf.text if closest_entity.dxftype() == 'TEXT' else closest_entity.plain_text()
            self.text_clicked.emit(text.strip())


class MessprotokollWidget(QWidget):
    # ... (unverändert) ...
    selection_mode_changed = Signal(str)
    field_selected = Signal(object)
    field_manually_edited = Signal(object)
    TOTAL_MEASURES = 18
    MEASURES_PER_PAGE = 6
    BLOCKS_PER_PAGE = 2
    MEASURES_PER_BLOCK = 3
    TOTAL_BLOCKS = TOTAL_MEASURES // MEASURES_PER_BLOCK
    _pos_vals = [f"+{i / 1000.0:.3f}" for i in range(5, 201, 5)]
    _neg_vals = [f"-{i / 1000.0:.3f}" for i in range(5, 201, 5)]
    TOLERANCE_VALUES = ["", "0"] + _neg_vals[::-1] + _pos_vals
    MESSMITTEL_OPTIONS = ["", "Aussen Mikrometer", "Digimar", "Endmaß", "Gewinde-lehrdorn", "Gewinde-lehrring",
                          "Haarlineal", "Innen Mikrometer", "Innenschnell-taster", "Lehrdorn", "Lehrring",
                          "MahrSurf M 310", "Maschinen- taster", "Mess-schieber", "Messuhr", "optisch", "Prüfstifte",
                          "Radius Lehre", "Rugotest", "Steigungs-lehre", "Subito", "Tiefenmaß", "Winkel-messer",
                          "Zeiss", "Zoller"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.cell_mapping = {}
        self._load_mapping()
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)
        data_dir = resource_path("Data")
        self.iso_calculator = IsoFitsCalculator(data_dir)
        self.nominal_fields, self.upper_tol_combos, self.lower_tol_combos = [], [], []
        self.soll_labels, self.iso_fit_combos, self.messmittel_combos = [], [], []
        self.measure_blocks = []
        self.current_page = 0
        self.total_pages = math.ceil(self.TOTAL_BLOCKS / self.BLOCKS_PER_PAGE)
        self._create_header(main_layout)
        main_layout.addSpacing(15)
        self._create_measure_blocks(main_layout)
        main_layout.addStretch()
        self._create_footer_controls(main_layout)
        self._update_page_view()

    def _create_header(self, main_layout):
        header_layout = QHBoxLayout()
        header_layout.setSpacing(30)
        fields_grid = QGridLayout()
        fields_grid.setVerticalSpacing(15)
        fields_grid.setHorizontalSpacing(10)
        title_label = QLabel("Messprotokoll-Assistent")
        title_label.setStyleSheet("font-size: 28pt; font-weight: bold;")
        fields_grid.addWidget(title_label, 0, 0, 1, 4)
        drawing_num_layout = QHBoxLayout()
        drawing_num_layout.addWidget(QLabel("Zeichnungsnummer:"))
        self.zeichnungsnummer_field = ClickableLineEdit()
        self.zeichnungsnummer_field.clicked.connect(self._on_zeichnungsnummer_field_selected)
        drawing_num_layout.addWidget(self.zeichnungsnummer_field, 1)
        fields_grid.addLayout(drawing_num_layout, 1, 0, 1, 4)
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(10)
        auftrag_label = QLabel("Auftrag:")
        self.auftrag_edit = QLineEdit()
        row2_layout.addWidget(auftrag_label)
        row2_layout.addWidget(self.auftrag_edit, 1)
        row2_layout.addSpacing(20)
        row2_layout.addWidget(QLabel("Pos.:"))
        self.pos_edit = QLineEdit()
        self.pos_edit.setFixedWidth(80)
        row2_layout.addWidget(self.pos_edit)
        row2_layout.addSpacing(20)
        row2_layout.addWidget(QLabel("Datum:"))
        self.date_edit = QDateEdit(calendarPopup=True, date=QDate.currentDate())
        row2_layout.addWidget(self.date_edit)
        fields_grid.addLayout(row2_layout, 2, 0, 1, 4)
        fields_grid.addWidget(QLabel("Oberflächenbehandlung:"), 3, 0, alignment=Qt.AlignTop)
        self.oberflaeche_edit = QLineEdit()
        fields_grid.addWidget(self.oberflaeche_edit, 3, 1, 1, 3)
        fields_grid.addWidget(QLabel("Bemerkungen:"), 4, 0, alignment=Qt.AlignTop)
        self.bemerkungen_edit = QLineEdit()
        fields_grid.addWidget(self.bemerkungen_edit, 4, 1, 1, 3)
        header_layout.addLayout(fields_grid, 1)
        logo_label = QLabel()
        logo_label.setFixedSize(200, 200)
        logo_label.setScaledContents(True)
        logo_path = resource_path("assets/logo.png")
        if os.path.exists(logo_path):
            logo_label.setPixmap(QPixmap(logo_path))
        header_layout.addWidget(logo_label, 0, alignment=Qt.AlignTop | Qt.AlignRight)
        main_layout.addLayout(header_layout)

    def get_scale_factor(self) -> float:
        text = self.scale_combo.currentText().strip().replace(',', '.')
        if ":" in text:
            try:
                parts = text.split(":")
                if len(parts) == 2:
                    return float(parts[0]) / float(parts[1])
            except (ValueError, ZeroDivisionError):
                return 1.0
        else:
            try:
                return float(text)
            except ValueError:
                return 1.0
        return 1.0

    def _create_footer_controls(self, main_layout):
        pagination_layout = QHBoxLayout()
        self.prev_button = QPushButton("<< Zurück")
        self.prev_button.clicked.connect(self._previous_page)
        self.page_label = QLabel("", alignment=Qt.AlignCenter)
        self.next_button = QPushButton("Vor >>")
        self.next_button.clicked.connect(self._next_page)
        pagination_layout.addStretch()
        pagination_layout.addWidget(self.prev_button)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.next_button)
        pagination_layout.addStretch()
        scale_layout = QHBoxLayout()
        scale_label = QLabel("Maßstab:")
        scale_label.setStyleSheet("font-weight: bold;")
        scale_layout.addWidget(scale_label)
        self.scale_combo = QComboBox()
        self.scale_combo.setEditable(True)
        self.scale_combo.setFixedWidth(100)
        self.scale_combo.addItems([
            "1:1", "2:1", "1:2", "5:1", "1:5", "10:1", "1:10", "1:20", "1:50", "1:100"
        ])
        scale_layout.addWidget(self.scale_combo)
        self.load_button = QPushButton("Protokoll Laden")
        self.load_button.clicked.connect(self._load_protokoll_from_excel)
        self.save_button = QPushButton("Protokoll Speichern")
        self.save_button.setProperty('class', 'success-color')
        self.save_button.clicked.connect(self._save_protokoll)
        bottom_layout = QHBoxLayout()
        bottom_layout.addLayout(pagination_layout)
        bottom_layout.addStretch(1)
        bottom_layout.addLayout(scale_layout)
        bottom_layout.addStretch(1)
        bottom_layout.addWidget(self.load_button)
        bottom_layout.addWidget(self.save_button)
        main_layout.addLayout(bottom_layout)

    def _create_measure_blocks(self, main_layout):
        for block_idx in range(self.TOTAL_BLOCKS):
            block_frame = QFrame()
            grid = QGridLayout(block_frame)
            grid.setSpacing(10)
            grid.addWidget(QLabel("Maß lt.\nZeichnung"), 0, 0, 7, 1)
            soll_qlabel = QLabel("SOLL ➡")
            soll_qlabel.setStyleSheet("font-weight: bold;")
            grid.addWidget(soll_qlabel, 6, 0, 1, 1)
            for col_idx in range(self.MEASURES_PER_BLOCK):
                idx = block_idx * self.MEASURES_PER_BLOCK + col_idx
                col = 1 + col_idx
                grid.addWidget(QLabel(f"Maß {idx + 1}", alignment=Qt.AlignCenter), 0, col)
                nf = ClickableLineEdit(alignment=Qt.AlignCenter)
                nf.clicked.connect(lambda w=nf: self._on_measure_field_selected(w))
                nf.textEdited.connect(lambda t, w=nf: self.field_manually_edited.emit(w))
                grid.addWidget(nf, 1, col);
                self.nominal_fields.append(nf)
                grid.addWidget(QLabel("ISO-Toleranz"), 2, col, alignment=Qt.AlignCenter)
                ifc = QComboBox();
                ifc.setEditable(True);
                ifc.addItems(self.iso_calculator.available_fits)
                grid.addWidget(ifc, 3, col);
                self.iso_fit_combos.append(ifc)
                grid.addWidget(QLabel("Messmittel"), 4, col, alignment=Qt.AlignCenter)
                mc = QComboBox();
                mc.addItems(self.MESSMITTEL_OPTIONS)
                grid.addWidget(mc, 5, col);
                self.messmittel_combos.append(mc)
                tol_layout = QGridLayout()
                utc = QComboBox();
                utc.addItems(self.TOLERANCE_VALUES);
                utc.setEditable(True)
                ltc = QComboBox();
                ltc.addItems(self.TOLERANCE_VALUES);
                ltc.setEditable(True)
                tol_layout.addWidget(utc, 0, 0);
                tol_layout.addWidget(QLabel("Größtmaß"), 0, 1)
                tol_layout.addWidget(ltc, 1, 0);
                tol_layout.addWidget(QLabel("Kleinstmaß"), 1, 1)
                self.upper_tol_combos.append(utc);
                self.lower_tol_combos.append(ltc)
                sl = QLabel("---", alignment=Qt.AlignCenter,
                            styleSheet="font-weight: bold; border: 1px solid grey; padding: 6px;")
                grid.addLayout(tol_layout, 6, col);
                grid.addWidget(sl, 7, col);
                self.soll_labels.append(sl)
                nf.textChanged.connect(lambda _, i=idx: self._trigger_iso_fit_calculation(i))
                ifc.currentTextChanged.connect(lambda _, i=idx: self._trigger_iso_fit_calculation(i))
                utc.currentTextChanged.connect(lambda _, i=idx: self._update_soll_wert(i))
                ltc.currentTextChanged.connect(lambda _, i=idx: self._update_soll_wert(i))
            main_layout.addWidget(block_frame)
            self.measure_blocks.append(block_frame)

    def _on_zeichnungsnummer_field_selected(self):
        self.field_selected.emit(self.zeichnungsnummer_field)
        self.selection_mode_changed.emit('text')

    def _on_measure_field_selected(self, widget):
        self.field_selected.emit(widget)
        self.selection_mode_changed.emit('dimension')

    def _load_mapping(self):
        try:
            mapping_file_path = resource_path("mapping.json")
            with open(mapping_file_path, 'r', encoding='utf-8') as f:
                self.cell_mapping = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            QMessageBox.critical(self, "Mapping Fehler",
                                 f"Die Datei 'mapping.json' konnte nicht geladen werden.\n\n{e}\n\nDie Speicherfunktion ist deaktiviert.")
            self.cell_mapping = {}

    def _get_cell_value(self, sheet, cell_coord):
        cell = sheet[cell_coord]
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return cell.value

    def _clear_ui(self):
        self.zeichnungsnummer_field.clear()
        self.auftrag_edit.clear()
        self.pos_edit.clear()
        self.oberflaeche_edit.clear()
        self.bemerkungen_edit.clear()
        self.date_edit.setDate(QDate.currentDate())
        for i in range(self.TOTAL_MEASURES):
            self.nominal_fields[i].clear()
            self.iso_fit_combos[i].setCurrentIndex(0)
            self.messmittel_combos[i].setCurrentIndex(0)
            self.upper_tol_combos[i].setCurrentIndex(0)
            self.lower_tol_combos[i].setCurrentIndex(0)
            self.soll_labels[i].setText("---")

    def _load_protokoll_from_excel(self):
        if not self.cell_mapping:
            QMessageBox.warning(self, "Aktion nicht möglich", "Die Mapping-Datei ist fehlerhaft oder fehlt.")
            return
        open_path, _ = QFileDialog.getOpenFileName(self, "Protokoll laden...", "", "Excel-Dateien (*.xlsx *.xls)")
        if not open_path:
            return
        try:
            self._clear_ui()
            workbook = openpyxl.load_workbook(open_path, data_only=True)
            sheet = workbook["Tabelle1"]
            header_map = self.cell_mapping.get("header", {})
            for key, cell_coord in header_map.items():
                value = self._get_cell_value(sheet, cell_coord)
                if value is None: continue
                value_str = str(value).strip()
                if key == 'zeichnungsnummer':
                    self.zeichnungsnummer_field.setText(value_str)
                elif key == 'auftrag':
                    self.auftrag_edit.setText(value_str)
                elif key == 'position':
                    self.pos_edit.setText(value_str)
                elif key == 'datum':
                    if isinstance(value, QDate):
                        self.date_edit.setDate(value)
                    else:
                        self.date_edit.setDate(QDate.fromString(value_str, "dd.MM.yyyy"))
                elif key == 'oberflaeche':
                    self.oberflaeche_edit.setText(value_str.replace("Oberflächenbehandlung:", "").strip())
                elif key == 'bemerkungen':
                    self.bemerkungen_edit.setText(value_str.replace("Bemerkungen:", "").strip())
            measure_map = self.cell_mapping.get("measures", [])
            for i in range(min(self.TOTAL_MEASURES, len(measure_map))):
                cell_info = measure_map[i]
                for key, cell_coord in cell_info.items():
                    value = self._get_cell_value(sheet, cell_coord)
                    if value is None: continue
                    value_str = str(value)
                    if key == 'nominal':
                        self.nominal_fields[i].setText(value_str.replace('.', ','))
                    elif key == 'iso_fit':
                        self.iso_fit_combos[i].setCurrentText(value_str)
                    elif key == 'messmittel':
                        self.messmittel_combos[i].setCurrentText(value_str)
                    elif key == 'upper_tol':
                        self.upper_tol_combos[i].setCurrentText(value_str)
                    elif key == 'lower_tol':
                        self.lower_tol_combos[i].setCurrentText(value_str)
            QMessageBox.information(self, "Erfolg", f"Protokoll '{os.path.basename(open_path)}' erfolgreich geladen.")
        except FileNotFoundError:
            QMessageBox.critical(self, "Fehler", f"Die Datei '{open_path}' wurde nicht gefunden.")
        except KeyError as e:
            QMessageBox.critical(self, "Excel-Fehler",
                                 f"Ein Schlüssel im Mapping ('{e}') oder das Arbeitsblatt 'Tabelle1' wurde nicht gefunden.")
        except Exception as e:
            QMessageBox.critical(self, "Ladefehler", f"Ein unerwarteter Fehler ist aufgetreten:\n\n{e}")

    def _save_protokoll(self):
        if not self.cell_mapping:
            QMessageBox.warning(self, "Speichern nicht möglich", "Die Mapping-Datei ist fehlerhaft oder fehlt.")
            return
        znr = self.zeichnungsnummer_field.text().strip()
        auftrag = self.auftrag_edit.text().strip()
        pos = self.pos_edit.text().strip()
        if not all([znr, auftrag, pos]):
            QMessageBox.warning(self, "Fehlende Eingabe", "Bitte 'Zeichnungsnummer', 'Auftrag' und 'Pos' ausfüllen.")
            return
        template_path = resource_path("LEERFORMULAR.xlsx")
        if not os.path.exists(template_path):
            QMessageBox.critical(self, "Vorlage fehlt", f"Die Vorlagendatei '{template_path}' wurde nicht gefunden.")
            return
        suggested_filename = f"{znr}+{auftrag}+{pos}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, "Protokoll speichern unter...", suggested_filename,
                                                   "Excel-Dateien (*.xlsx)")
        if not save_path: return
        try:
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook["Tabelle1"]
            header_map = self.cell_mapping.get("header", {})

            def write_cell(coord, value):
                cell = sheet[coord]
                if isinstance(cell, MergedCell):
                    m_range = [r for r in sheet.merged_cells.ranges if coord in r][0]
                    sheet.cell(row=m_range.min_row, column=m_range.min_col).value = value
                else:
                    cell.value = value

            if 'zeichnungsnummer' in header_map: write_cell(header_map['zeichnungsnummer'], znr)
            if 'auftrag' in header_map: write_cell(header_map['auftrag'], auftrag)
            if 'position' in header_map: write_cell(header_map['position'], pos)
            if 'datum' in header_map: write_cell(header_map['datum'], self.date_edit.date().toString("dd.MM.yyyy"))
            oberflaeche = self.oberflaeche_edit.text().strip()
            if oberflaeche and 'oberflaeche' in header_map: write_cell(header_map['oberflaeche'],
                                                                       f"Oberflächenbehandlung: {oberflaeche}")
            bemerkungen = self.bemerkungen_edit.text().strip()
            if bemerkungen and 'bemerkungen' in header_map: write_cell(header_map['bemerkungen'],
                                                                       f"Bemerkungen: {bemerkungen}")
            measure_map = self.cell_mapping.get("measures", [])
            for i in range(min(self.TOTAL_MEASURES, len(measure_map))):
                cell_info = measure_map[i]
                if 'nominal' in cell_info: write_cell(cell_info['nominal'], self.nominal_fields[i].text())
                if 'iso_fit' in cell_info: write_cell(cell_info['iso_fit'], self.iso_fit_combos[i].currentText())
                if 'messmittel' in cell_info: write_cell(cell_info['messmittel'],
                                                         self.messmittel_combos[i].currentText())
                if 'upper_tol' in cell_info: write_cell(cell_info['upper_tol'], self.upper_tol_combos[i].currentText())
                if 'lower_tol' in cell_info: write_cell(cell_info['lower_tol'], self.lower_tol_combos[i].currentText())
                if 'soll' in cell_info: write_cell(cell_info['soll'], self.soll_labels[i].text())
            workbook.save(filename=save_path)
            QMessageBox.information(self, "Erfolg", f"Protokoll erfolgreich unter '{save_path}' gespeichert.")
        except Exception as e:
            QMessageBox.critical(self, "Speicherfehler", f"Ein unerwarteter Fehler ist aufgetreten:\n\n{e}")

    def _update_page_view(self):
        start_block = self.current_page * self.BLOCKS_PER_PAGE
        end_block = start_block + self.BLOCKS_PER_PAGE
        for i, block in enumerate(self.measure_blocks):
            block.setVisible(start_block <= i < end_block)
        self.page_label.setText(f"Seite {self.current_page + 1} / {self.total_pages}")
        self.prev_button.setEnabled(self.current_page > 0)
        self.next_button.setEnabled(self.current_page < self.total_pages - 1)

    def _previous_page(self):
        if self.current_page > 0: self.current_page -= 1; self._update_page_view()

    def _next_page(self):
        if self.current_page < self.total_pages - 1: self.current_page += 1; self._update_page_view()

    def _trigger_iso_fit_calculation(self, index):
        self._update_soll_wert(index)
        nominal_text = self.nominal_fields[index].text().replace(',', '.')
        fit_string = self.iso_fit_combos[index].currentText().strip()
        if not nominal_text or not fit_string: return
        try:
            result = self.iso_calculator.calculate(float(nominal_text), fit_string)
            if result:
                up_dev, low_dev = result
                self.upper_tol_combos[index].blockSignals(True)
                self.lower_tol_combos[index].blockSignals(True)
                self.upper_tol_combos[index].setCurrentText(f"{up_dev:+.3f}")
                self.lower_tol_combos[index].setCurrentText(f"{low_dev:+.3f}")
                self.upper_tol_combos[index].blockSignals(False)
                self.lower_tol_combos[index].blockSignals(False)
                self._update_soll_wert(index)
        except (ValueError, TypeError):
            pass

    def _update_soll_wert(self, index):
        try:
            nominal = float(self.nominal_fields[index].text().replace(',', '.') or 0)
            upper_tol = float(self.upper_tol_combos[index].currentText().replace(',', '.') or 0)
            lower_tol = float(self.lower_tol_combos[index].currentText().replace(',', '.') or 0)
            soll_wert = nominal + (upper_tol + lower_tol) / 2.0
            self.soll_labels[index].setText(f"{soll_wert:.4f}".replace('.', ','))
        except (ValueError, TypeError):
            self.soll_labels[index].setText("---")


# ==============================================================================
#      4. HAUPTFENSTER (MAIN WINDOW)
# ==============================================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.target_widget = None
        app_icon_path = resource_path("assets/logo.png")
        self.setWindowIcon(QIcon(app_icon_path))
        self._set_application_style()
        self.setAcceptDrops(True)
        self.dxf_widget = DXFWidget()
        self.protokoll_widget = MessprotokollWidget()
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.addWidget(self.dxf_widget)
        self.splitter.addWidget(self.protokoll_widget)
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 1)
        self.splitter.handle(1).setEnabled(False)
        self.splitter.setHandleWidth(1)
        self.splitter.setChildrenCollapsible(False)
        self.setCentralWidget(self.splitter)

        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("Datei")
        open_action = QAction("DXF Öffnen...", self)
        open_action.triggered.connect(self.open_file)
        file_menu.addAction(open_action)
        self.setWindowTitle("Messprotokoll-Assistent")
        self.setMinimumSize(1024, 768)
        screen_geometry = QApplication.primaryScreen().availableGeometry()
        self.resize(int(screen_geometry.width() * 0.9), int(screen_geometry.height() * 0.9))
        self.move(int(screen_geometry.width() * 0.05), int(screen_geometry.height() * 0.05))
        total_width = self.width()
        half_width = total_width // 2
        self.splitter.setSizes([half_width, half_width])
        self.protokoll_widget.selection_mode_changed.connect(self.dxf_widget.set_selection_mode)
        self.protokoll_widget.field_selected.connect(self.on_field_selected)
        self.dxf_widget.dimension_clicked.connect(self.on_dimension_value_received)
        self.dxf_widget.text_clicked.connect(self.on_text_value_received)
        self.protokoll_widget.field_manually_edited.connect(self.on_field_manually_edited)

    def _set_application_style(self):
        app = QApplication.instance()
        extra = {'accent_color': '#448AFF', 'secondaryLightColor': "#31363B"}
        apply_stylesheet(app, theme='dark_blue.xml', extra=extra)

    def on_field_selected(self, widget):
        if self.target_widget: self.target_widget.setStyleSheet("")
        self.target_widget = widget
        self.target_widget.setStyleSheet("border: 2px solid #448AFF;")

    def on_dimension_value_received(self, value_str):
        if self.target_widget:
            try:
                drawn_value = float(value_str.replace(',', '.'))
                scale_factor = self.protokoll_widget.get_scale_factor()

                if scale_factor > 0:
                    real_value = drawn_value / scale_factor
                else:
                    real_value = drawn_value

                corrected_value_str = f"{real_value:.4f}".replace('.', ',')
                self.target_widget.setText(corrected_value_str)

            except (ValueError, TypeError):
                self.target_widget.setText(value_str.replace('.', ','))

            self.target_widget.setStyleSheet("")
            self.target_widget = None
        else:
            QMessageBox.information(self, "Hinweis", "Bitte zuerst in ein 'Maß lt. Zeichnung'-Feld klicken.")

    def on_text_value_received(self, value):
        if self.target_widget:
            self.target_widget.setText(value)
            self.target_widget.setStyleSheet("")
            self.target_widget = None
        else:
            QMessageBox.information(self, "Hinweis", "Bitte zuerst in das 'Zeichnungsnummer'-Feld klicken.")

    def on_field_manually_edited(self, widget):
        if self.target_widget == widget:
            self.target_widget.setStyleSheet("")
            self.target_widget = None

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "DXF-Datei öffnen", "", "DXF-Dateien (*.dxf)")
        if path: self.dxf_widget.load_dxf(path)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls() and event.mimeData().urls()[0].isLocalFile():
            path = event.mimeData().urls()[0].toLocalFile()
            if path.lower().endswith('.dxf'):
                event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        path = event.mimeData().urls()[0].toLocalFile()
        self.dxf_widget.load_dxf(path)


# ==============================================================================
#      5. START DER APPLIKATION
# ==============================================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())